import json
import pandas as pd
import os
from openpyxl import load_workbook


def read_jsonl(jsonl_file):
    with open(jsonl_file, "r", encoding="utf-8") as f:
        lines = []
        for i, line in enumerate(f):
            lines.append(json.loads(line.strip()))  # strip()用于去掉两边多余空格
        return lines


class CellRel(object):
    def __init__(self, details, rel):
        self.details = details
        self.rel = rel


def get_label(cell_str):
    if cell_str is None:
        return "0", cell_str
    if len(cell_str) > 1 and cell_str[-2] == '_':
        if cell_str[-1] == "0" or cell_str[-1] == "1" or cell_str[-1] == "2":
            label = cell_str[-1]
        else:
            label = "0"
        cell_str = cell_str[:-2]
    else:
        label = "0"
    return label, cell_str


def convert_xlsx_to_json(xlsx_file, json_input_file, json_output_file):
    # 加载文件
    book = load_workbook(xlsx_file)
    # sheet name获取sheet：
    worksheets = book.sheetnames
    json_input_data = read_jsonl(json_input_file)
    fw = open(json_output_file, 'w')
    for idx, worksheet in enumerate(worksheets):
        sheet = book[worksheet]
        seq = worksheet.split("_")[-1]
        # 获取总行数
        rows = sheet.max_row
        # 获取总列数
        # cols = sheet.max_column
        # print(rows)
        # 获取表头
        head = ([row for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True)][0])[1:]
        head_rel = []
        for head_cell in head:
            label, head_str = get_label(head_cell)
            # head_rel.append(json.dumps(CellRel(head_str, label)))
            if head_str is not None and len(head_str) >= 4 and head_str[0:4] == "NULL":
                head_str = ""
            head_rel.append({'details': head_str, 'rel': label})
        # head_rel = json.dumps(head_rel)
        # 数据组装
        data_rel = []
        for row in sheet.iter_rows(min_row=2, max_row=rows, values_only=True):
            row = row[1:]
            row_rel = []
            for cell in row:
                label, cell_str = get_label(str(cell))
                # row_rel.append(json.dumps(CellRel(cell_str, label)))
                row_rel.append({'details': cell_str, 'rel': label})
            data_rel.append(row_rel)
        # data_rel = json.dumps(data_rel)
        qid = json_input_data[idx]['qid']
        docid = json_input_data[idx]['docid']
        query = json_input_data[idx]['query']
        # caption = json_input_data[idx]['caption']
        rel = json_input_data[idx]['rel']
        num_cols = len(head_rel)
        num_rows = len(data_rel)
        dic = {'qid': qid, 'query': query, 'docid': docid, 'rel': rel,
               'num_cols': num_cols, 'num_rows': num_rows, 'head': head_rel, 'data': data_rel, 'seq': seq}
        json.dump(dic, fw)
        fw.write('\n')
    fw.close()


def main():
    # input_data = "../tmp/all_highly_relevant.xlsx"
    # convert_xlsx_to_json(input_data, "../data/all_highly_relevant.json", "../tmp/all_cell_rel.json")
    # input_data = "../tmp/all_highly_relevant.xlsx"
    # convert_xlsx_to_json(input_data, "../data/all_highly_relevant.json", "../tmp/all_cell_rel2.json")
    input_data = "../tmp/wdc_all_highly_relevant.xlsx"
    convert_xlsx_to_json(input_data, "../data/wdc_all_highly_rekevant.jsonl", "../tmp/wdc_all_cell_rel.json")


if __name__ == '__main__':
    main()
